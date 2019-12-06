#!/usr/bin/env python3

"""
Required fields for the manifest:
* sampleid - a string found in the fastq file name uniquely identifying a
  specimen
* batch - a label grouping specimens into PCR batches for dada2::learnErrors()

Verifies the following:
* all sampleids are unique
* every sampleid in the manifest has corresponding R{1,2} and I{1,2}
* all fastq file names have sampleid as the first underscore-delimited token

"""

import argparse
import csv
import glob
import itertools
import operator
import os
import sys
from collections import defaultdict
import re

import openpyxl

KEEPCOLS = {'sampleid', 'sample_name', 'project', 'batch', 'controls'}


def read_manifest_excel(fname, keepcols=KEEPCOLS):
    """Read the first worksheet from an excel file and return a generator
    of dicts with keys limited to 'keepcols'.

    """

    valgetter = operator.attrgetter('value')

    wb = openpyxl.load_workbook(fname)
    sheet = wb[wb.sheetnames[0]]

    rows = (r for r in sheet.iter_rows() if r[0].value)

    # get fieldnames from cells in the first row up to the first empty one
    header = itertools.takewhile(valgetter, next(rows))
    # replace column names to be discarded with '_'
    fieldnames = [(cell.value if cell.value in keepcols else '_')
                  for cell in header]
    popextra = '_' in fieldnames

    for row in rows:
        d = dict(zip(fieldnames, map(valgetter, row)))
        if popextra:
            d.pop('_')
        if d['sampleid']:
            yield d


def read_manifest_csv(fname, keepcols=KEEPCOLS):

    with open(fname) as f:
        reader = csv.DictReader(f)
        reader.fieldnames = [(n if n in keepcols else '_')
                             for n in reader.fieldnames]
        popextra = '_' in reader.fieldnames
        for d in reader:
            if popextra:
                d.pop('_')
            if d['sampleid']:
                yield dict(d)


def get_sampleid(pth):
    return os.path.basename(pth).split('_')[0]


def main(arguments):
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('manifest', help="Manifest in excel or csv format")
    parser.add_argument('fastq_files', type=argparse.FileType('r'),
                        help="File listing fastq inputs")
    parser.add_argument('-o', '--outfile', help="Output .json file",
                        type=argparse.FileType('w'), default=sys.stdout)
    args = parser.parse_args(arguments)

    if args.manifest.endswith('.csv'):
        read_manifest = read_manifest_csv
    else:
        read_manifest = read_manifest_excel

    manifest = list(read_manifest(args.manifest))
    manifest_sampleids = {row['sampleid'] for row in manifest}

    fq_files = sorted(line.strip() for line in args.fastq_files if line.strip())

    fq_sampleids = {get_sampleid(pth) for pth in fq_files}

    # make sure all sampleids are unique
    assert len(manifest) == len(manifest_sampleids)

    # confirm that all sampleids in the manifest have corresponding
    # fastq files
    extras = manifest_sampleids - fq_sampleids
    if extras:
        sys.exit('samples in the manifest without fastq files: {}'.format(extras))

    # confirm that every sampleid is represented by four fastq files
    for sampleid, paths in itertools.groupby(fq_files, key=get_sampleid):
        labels = [re.findall(r'_([IR][12])_', fname)[0] for fname in paths]
        if labels != ['I1', 'I2', 'R1', 'R2']:
            sys.exit('a fastq file missing for sampleid {}: has {}'.format(
                sampleid, labels))

    # finally, write an output file with columns (sampleid, batch)
    writer = csv.DictWriter(
        args.outfile, fieldnames=['sampleid', 'batch'], extrasaction='ignore')
    writer.writeheader()
    writer.writerows(manifest)


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
